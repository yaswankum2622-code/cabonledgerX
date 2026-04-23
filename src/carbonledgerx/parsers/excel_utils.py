"""Excel-specific helpers for workbook inspection."""

from __future__ import annotations

from collections import Counter
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
import re

import openpyxl
import pandas as pd


HEADER_SCAN_LIMIT = 50
EMPTY_HEADER_FALLBACK = "unnamed"


@dataclass(frozen=True, slots=True)
class SheetLoadResult:
    """Container for a loaded worksheet and its detected header row."""

    sheet_name: str
    header_row_number: int
    dataframe: pd.DataFrame


def list_excel_sheet_names(file_path: str | Path) -> list[str]:
    """Return workbook sheet names in file order."""

    workbook_path = Path(file_path)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def trim_header_whitespace(headers: Sequence[object]) -> list[str]:
    """Trim and normalize internal whitespace for header values."""

    cleaned_headers: list[str] = []
    for value in headers:
        if value is None or pd.isna(value):
            cleaned_headers.append("")
            continue

        text = str(value).strip()
        text = re.sub(r"\s+", " ", text)
        cleaned_headers.append(text)

    return cleaned_headers


def normalize_column_name(column_name: object) -> str:
    """Convert a raw column name to a snake_case identifier."""

    header = trim_header_whitespace([column_name])[0]
    if not header:
        return EMPTY_HEADER_FALLBACK

    normalized = header.lower()
    normalized = normalized.replace("%", " percent ")
    normalized = re.sub(r"[^0-9a-z]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")

    if not normalized:
        return EMPTY_HEADER_FALLBACK
    if normalized[0].isdigit():
        return f"col_{normalized}"

    return normalized


def make_unique_column_names(column_names: Sequence[str]) -> list[str]:
    """Ensure column names remain unique while preserving order."""

    counts: Counter[str] = Counter()
    unique_names: list[str] = []

    for name in column_names:
        base_name = name or EMPTY_HEADER_FALLBACK
        counts[base_name] += 1

        if counts[base_name] == 1:
            unique_names.append(base_name)
            continue

        unique_names.append(f"{base_name}_{counts[base_name]}")

    return unique_names


def normalize_column_names(headers: Sequence[object]) -> list[str]:
    """Trim, normalize, and de-duplicate a header sequence."""

    normalized = [normalize_column_name(header) for header in headers]
    return make_unique_column_names(normalized)


def standardize_dataframe_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of a dataframe with normalized column names."""

    standardized = dataframe.copy()
    standardized.columns = normalize_column_names(list(standardized.columns))
    return standardized


def detect_header_row(
    file_path: str | Path,
    sheet_name: str,
    *,
    scan_limit: int = HEADER_SCAN_LIMIT,
) -> int:
    """Detect the most plausible header row in the top portion of a worksheet."""

    workbook_path = Path(file_path)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]

    best_row_number = 1
    best_score = float("-inf")

    try:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
            start=1,
        ):
            score = _score_header_candidate(row)
            if score > best_score:
                best_score = score
                best_row_number = row_number
    finally:
        workbook.close()

    return best_row_number


def read_excel_sheet_safely(
    file_path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int | None = None,
    normalize_headers: bool = False,
) -> SheetLoadResult:
    """Load a worksheet into pandas using a detected or supplied header row."""

    workbook_path = Path(file_path)
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected an .xlsx workbook, received: {workbook_path}")

    resolved_header_row = header_row_number or detect_header_row(workbook_path, sheet_name)
    dataframe = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=resolved_header_row - 1,
        engine="openpyxl",
    )
    dataframe = dataframe.dropna(axis="rows", how="all").dropna(axis="columns", how="all")

    if normalize_headers:
        dataframe = standardize_dataframe_columns(dataframe)

    return SheetLoadResult(
        sheet_name=sheet_name,
        header_row_number=resolved_header_row,
        dataframe=dataframe,
    )


def _score_header_candidate(row_values: Sequence[object]) -> float:
    """Score a row for header-likeness using lightweight heuristics."""

    non_empty_values = [value for value in row_values if not _is_empty_cell(value)]
    if not non_empty_values:
        return float("-inf")

    string_values = [str(value).strip() for value in non_empty_values]
    non_empty_count = len(string_values)
    unique_ratio = len(set(string_values)) / non_empty_count
    readable_count = sum(_is_readable_header_text(text) for text in string_values)
    code_like_count = sum(_is_code_like(text) for text in string_values)
    colon_count = sum(text.endswith(":") for text in string_values)
    long_text_count = sum(len(text) > 80 for text in string_values)
    numeric_like_count = sum(_is_numeric_like(value) for value in non_empty_values)

    score = (
        non_empty_count
        + (unique_ratio * 5.0)
        + (readable_count * 0.5)
        - (code_like_count * 0.25)
        - (colon_count * 1.5)
        - (long_text_count * 2.0)
        - (numeric_like_count * 0.75)
    )

    if non_empty_count == 1:
        score -= 10.0
    elif non_empty_count == 2:
        score -= 3.0

    return score


def _is_empty_cell(value: object) -> bool:
    """Return whether a worksheet cell should be treated as empty."""

    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()

    return bool(pd.isna(value))


def _is_readable_header_text(text: str) -> bool:
    """Return whether a string looks like a human-readable header label."""

    return (" " in text) or any(character.islower() for character in text)


def _is_code_like(text: str) -> bool:
    """Return whether a cell looks like a compact code rather than a header."""

    return bool(re.fullmatch(r"[A-Z0-9_./()%+-]+", text))


def _is_numeric_like(value: object) -> bool:
    """Return whether a cell is numerically typed."""

    return isinstance(value, (int, float)) and not isinstance(value, bool)
